import os, uuid
import xlrd
from xlutils.copy import copy
import xlwt
import openpyxl
import pandas as pd
import pyexcel as p
from openpyxl.styles import NamedStyle
import re
import time 
import datetime
from dateutil.parser import parse
from azure.storage.blob import BlobServiceClient, BlobClient, ContainerClient, __version__


def uploadfiletoAzureBlob(file, filename):
    try:
        print("Azure Blob storage v" + __version__ + " - Python quickstart sample")
        connect_str = 'YOUR AZURE BLOB CONN STRING'
        blob_service_client = BlobServiceClient.from_connection_string(connect_str)
        container_name = 'YOUR CONTAINER NAME IN BLOB STORAGE'
        blob_client = blob_service_client.get_blob_client(container=container_name, blob=filename)

        print("\nUploading to Azure Storage as blob:\n\t" + file)

        with open(file, "rb") as data:
            blob_client.upload_blob(data)
    except Exception as ex:
        print('Exception:')
        print(ex)

#parse date from file name string
def getdatefromfilename(filename):
    new_string = re.findall('Name -(.*)$', filename)
    new_string = new_string[0].strip()
    new_string = new_string[0:7]
    conv=parse(new_string)
    new_string = conv.strftime('%m%d%Y')
    return new_string

#return interpolated string
def getnamestringusingcityanddate(city,date):
    filenamestring = f'Atlanta.xlsx'
    return filenamestring

#finds a city in your file name or returns Charlotte
def getcityfromfilename(filename):
    new_string = re.findall('Your test string(.*)$', filename)
    new_string = new_string[0].strip()
    new_string = new_string.replace(" ", "")
    new_string = new_string.replace(".xlsx", "")
    new_string = new_string.title()
    if(new_string == ""):
        return "Charlotte"
    else:    
        return new_string


path = 'C:\excelfiles'
targetdir = (path + "/New_Files/") #where you want your new files

if not os.path.exists(targetdir): #makes your new directory
    os.makedirs(targetdir)

for root,dir,files in os.walk(path, topdown=False): #all the files you want to split
    xlsfiles=[f for f in files] #can add selection condition here
for f in xlsfiles:
    wb = xlrd.open_workbook(os.path.join(root, f), on_demand=True)
    for sheet in wb.sheets(): #cycles through each sheet in each workbook
        newwb = copy(wb) #makes a temp copy of that book
        newwb._Workbook__worksheets = [ worksheet for worksheet in newwb._Workbook__worksheets if worksheet.name == sheet.name ]
        #brute force, but strips away all other sheets apart from the sheet being looked at
        namer = targetdir + f.strip(".xls") + sheet.name + ".xls"
        newwb.save(namer.replace(',','')) 
        #saves each sheet as the original file name plus the sheet name
     

path2 = 'C:/excelfiles/New_Files/'
for root,dir,files in os.walk(path2, topdown=False):
    xlsfiles2=[t for t in files]
#rename files to xlsx to interact with openpyxl
for p3 in xlsfiles2:
    new_stringer = getcityfromfilename(p3)
    datefromname= getdatefromfilename(p3)
    wholename = getnamestringusingcityanddate(new_stringer,datefromname)
    pathandfilename = path2 + p3
    pathandfilenamexls = pathandfilename.replace('.xls','.xlsx')
    p.save_book_as(file_name= pathandfilename,dest_file_name=pathandfilenamexls)
    os.remove(pathandfilename)



for root,dir,files in os.walk(path2, topdown=False):
    xlsfiles3=[d for d in files]
#delete columns and format cells
for p4 in xlsfiles3:
    filepathcomplete= path2 + p4
    book = openpyxl.load_workbook(filepathcomplete)
    sheenames = book.sheetnames[0]
    sheet = book[sheenames]
    sheet.delete_rows(1,8)
    sheet.delete_cols(11)
    sheet.delete_cols(5)
    date_style = NamedStyle(name='datetime', number_format='MM/DD/YYYY')
    for col in range(1,2):
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row,column=col).style = date_style
    for col in range(10,11):
        for row in range(2, sheet.max_row + 1):
            sheet.cell(row=row,column=col).number_format = '0.00'        
    book.save(filepathcomplete)  
    book.close()



for root,dir,files in os.walk(path2, topdown=False):
    xlsfiles3=[d for d in files]
#rename files using methods defined above then call function to upload to azure blob storage
for p5 in xlsfiles3:
    new_stringer = getcityfromfilename(p5)
    datefromname= getdatefromfilename(p5)
    wholename = getnamestringusingcityanddate(new_stringer,datefromname)
    pathandfilename = path2 + p5
    wholenameandpath = path2 + wholename
    os.rename(pathandfilename, wholenameandpath)
    uploadfiletoAzureBlob(wholenameandpath, wholename)

